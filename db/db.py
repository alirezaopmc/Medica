from openpyxl import load_workbook


def getUUIDS(path: str) -> list:
    wb = load_workbook(path)
    ws = wb.active

    return set(map(lambda cell: cell.value, ws['A']))
